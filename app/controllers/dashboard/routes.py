from flask import render_template, Blueprint, redirect, url_for, flash, request, current_app, send_file, abort
from flask_login import login_required, current_user
from app import db
from app.controllers.dashboard import dashboard_bp
from app.models.member import Member
from app.models.subscription import Subscription
from app.models.vehicle import Vehicle, Driver
from app.models.finance import Income, Expense, CashBox, BankAccount
from app.models.student import Student, TransportSubscription, EducationalLevel
from app.models.activity import Activity, Meeting
from datetime import datetime, timedelta, date
import calendar
import random
import json
from sqlalchemy import extract, func, and_
from app.models.about import About, BoardMember, Mandate
from app.forms.about import AboutForm, BoardMemberForm, MandateForm
import logging
import os
from werkzeug.utils import secure_filename
import csv
from io import BytesIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.views import SheetView

@dashboard_bp.route('/')
@dashboard_bp.route('/index')
@login_required
def index():
    # Statistics
    try:
        # Total counts
        total_students = Student.query.count()
        total_transport_subscriptions = TransportSubscription.query.count()
        
        # Current month stats
        today = date.today()
        current_month = today.month
        current_year = today.year
        current_month_name = calendar.month_name[current_month]
        current_month_subscriptions = TransportSubscription.query.filter(
            and_(
                extract('month', TransportSubscription.payment_date) == current_month,
                extract('year', TransportSubscription.payment_date) == current_year
            )
        ).count()
        
        # Calculate revenue
        current_month_revenue = TransportSubscription.query.filter(
            and_(
                extract('month', TransportSubscription.payment_date) == current_month,
                extract('year', TransportSubscription.payment_date) == current_year
            )
        ).with_entities(TransportSubscription.amount).all()
        total_revenue = sum([revenue[0] for revenue in current_month_revenue]) if current_month_revenue else 0
        
        # Calculate monthly stats for the past 12 months
        months = []
        member_subscriptions = []
        transport_subscriptions = []
        
        for i in range(12):
            month_date = today - timedelta(days=30*i)
            month = month_date.month
            year = month_date.year
            
            # Get month name in Arabic
            month_name = calendar.month_name[month]
            # Simplified month name
            short_month = f"{month_name[:3]} {str(year)[2:]}"
            months.append(short_month)
            
            # Member subscriptions (placeholder for now)
            member_count = 0
            member_subscriptions.append(member_count)
            
            # Transport subscriptions
            transport_count = TransportSubscription.query.filter(
                and_(
                    extract('month', TransportSubscription.payment_date) == month,
                    extract('year', TransportSubscription.payment_date) == year
                )
            ).count()
            transport_subscriptions.append(transport_count)
        
        # Reverse the lists to show oldest to newest
        months.reverse()
        member_subscriptions.reverse()
        transport_subscriptions.reverse()
        
        # Educational levels stats
        levels = EducationalLevel.query.order_by(EducationalLevel.order).all()
        level_names = [level.name for level in levels]
        level_counts = []
        
        for level in levels:
            count = Student.query.filter_by(educational_level=level.name).count()
            level_counts.append(count)
        
        # Colors for chart
        level_colors = [
            '#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', 
            '#858796', '#5a5c69', '#20c9a6', '#2c9faf', '#6610f2'
        ]
        # Ensure we have enough colors
        while len(level_colors) < len(levels):
            level_colors.extend(level_colors)
        level_colors = level_colors[:len(levels)]
        
        # Hover colors (slightly darker)
        level_hover_colors = [
            '#2e59d9', '#17a673', '#2c9faf', '#dda20a', '#be2617', 
            '#60616f', '#373840', '#169b7a', '#218a98', '#4d0bc4'
        ]
        while len(level_hover_colors) < len(levels):
            level_hover_colors.extend(level_hover_colors)
        level_hover_colors = level_hover_colors[:len(levels)]
        
        # Prepare stats for JSON
        stats = {
            'months': json.dumps(months),
            'monthly_member_subscriptions': json.dumps(member_subscriptions),
            'monthly_transport_subscriptions': json.dumps(transport_subscriptions),
            'level_names': json.dumps(level_names),
            'level_counts': json.dumps(level_counts),
            'level_colors': json.dumps(level_colors),
            'level_hover_colors': json.dumps(level_hover_colors)
        }
        
        return render_template('dashboard/index.html', 
                              title='لوحة التحكم',
                              total_students=total_students,
                              total_transport=total_transport_subscriptions,
                              current_month=current_month_name,
                              current_month_transport=current_month_subscriptions,
                              total_revenue=total_revenue,
                              stats=stats)
    except Exception as e:
        # Handle database errors or missing columns
        error_message = str(e)
        return render_template('dashboard/index.html', 
                             title='لوحة التحكم',
                             error=error_message,
                             total_students=0,
                             total_transport=0,
                             current_month="",
                             current_month_transport=0,
                             total_revenue=0,
                             stats={
                                 'months': '[]',
                                 'monthly_member_subscriptions': '[]',
                                 'monthly_transport_subscriptions': '[]',
                                 'level_names': '[]',
                                 'level_counts': '[]',
                                 'level_colors': '[]',
                                 'level_hover_colors': '[]'
                             })

@dashboard_bp.route('/about', methods=['GET', 'POST'])
@login_required
def about():
    about = About.query.first()
    
    # إذا لم يكن هناك كائن About، قم بإنشائه
    if not about:
        about = About(
            title='عن الجمعية',
            content='محتوى عن الجمعية'
        )
        db.session.add(about)
        db.session.commit()
    
    form = AboutForm(obj=about)
    if form.validate_on_submit() and current_user.is_admin:
        # رفع الشعار
        if form.logo.data:
            logo_filename = secure_filename(form.logo.data.filename)
            logo_path = os.path.join(current_app.root_path, 'static', 'uploads', 'about', logo_filename)
            os.makedirs(os.path.dirname(logo_path), exist_ok=True)
            form.logo.data.save(logo_path)
            about.logo = f'static/uploads/about/{logo_filename}'
        # رفع ملف القانون الأساسي
        if form.bylaws_file.data:
            bylaws_filename = secure_filename(form.bylaws_file.data.filename)
            bylaws_path = os.path.join(current_app.root_path, 'static', 'uploads', 'about', bylaws_filename)
            os.makedirs(os.path.dirname(bylaws_path), exist_ok=True)
            form.bylaws_file.data.save(bylaws_path)
            about.bylaws_file = f'static/uploads/about/{bylaws_filename}'
        # رفع الوصل القانوني النهائي
        if form.legal_doc.data:
            legal_doc_filename = secure_filename(form.legal_doc.data.filename)
            legal_doc_path = os.path.join(current_app.root_path, 'static', 'uploads', 'about', legal_doc_filename)
            os.makedirs(os.path.dirname(legal_doc_path), exist_ok=True)
            form.legal_doc.data.save(legal_doc_path)
            about.legal_doc = f'static/uploads/about/{legal_doc_filename}'
        # رفع صورة مقر الجمعية
        if form.hq_image.data:
            hq_image_filename = secure_filename(form.hq_image.data.filename)
            hq_image_path = os.path.join(current_app.root_path, 'static', 'uploads', 'about', hq_image_filename)
            os.makedirs(os.path.dirname(hq_image_path), exist_ok=True)
            form.hq_image.data.save(hq_image_path)
            about.hq_image = f'static/uploads/about/{hq_image_filename}'
        # الحقول النصية
        about.title = form.title.data
        about.content = form.content.data
        about.goals = form.goals.data
        about.principles = form.principles.data
        about.bylaws_text = form.bylaws_text.data
        db.session.commit()
        flash('تم تحديث بيانات الجمعية بنجاح.', 'success')
        return redirect(url_for('dashboard.about'))
    return render_template('dashboard/about.html', form=form, about=about)

@dashboard_bp.route('/about/edit', methods=['GET', 'POST'])
@login_required
def edit_about():
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('dashboard.about'))
    
    about = About.query.first()
    form = AboutForm()
    
    if form.validate_on_submit():
        if about:
            about.title = form.title.data
            about.content = form.content.data
        else:
            about = About(
                title=form.title.data,
                content=form.content.data
            )
            db.session.add(about)
        
        db.session.commit()
        flash('تم تحديث معلومات الجمعية بنجاح', 'success')
        return redirect(url_for('dashboard.about'))
    
    if about:
        form.title.data = about.title
        form.content.data = about.content
    
    return render_template('dashboard/forms/about_form.html', form=form, about=about)

@dashboard_bp.route('/board', methods=['GET', 'POST'])
@login_required
def board():
    print('==> دخلت دالة board')
    print('request.method:', request.method)
    print('request.form:', request.form)
    print('request.files:', request.files)
    members = BoardMember.query.all()
    mandates = Mandate.query.order_by(Mandate.start_date.desc()).all()
    form = BoardMemberForm()
    
    # التحقق من وجود فترات انتداب
    if not mandates:
        flash('يرجى إضافة فترة انتداب أولاً في <a href="{}" class="alert-link">قسم فترات الانتداب</a> قبل إضافة أعضاء المكتب.'.format(url_for('dashboard.mandate')), 'warning')
        form.mandate_id.choices = []
    else:
        form.mandate_id.choices = [(m.id, m.title) for m in mandates]
    
    if request.method == 'POST' and 'import_submit' in request.form and current_user.is_admin:
        print('==> دخلت شرط الاستيراد')
        flash('بدأت عملية الاستيراد...', 'info')
        try:
            print('==> قبل import openpyxl')
            from openpyxl import load_workbook
            print('==> بعد import openpyxl')
            print('==> قبل import pandas')
            import pandas as pd
            print('==> بعد import pandas')
            print('==> قبل import io')
            import io
            print('==> بعد import io')
            file = form.import_file.data
            print('==> بعد جلب الملف من الفورم')
            filename = file.filename
            print(f'اسم الملف: {filename}')
            imported_count = 0
            errors = []
            df = None
            file.stream.seek(0)
            print('بعد seek(0)')
            if filename.endswith('.xlsx'):
                print('سيتم قراءة الملف كـ Excel')
                in_memory_file = io.BytesIO(file.read())
                print('تم نسخ الملف في الذاكرة')
                wb = load_workbook(in_memory_file, data_only=True)
                print('تم فتح ملف Excel')
                ws = wb.active
                data = ws.values
                columns = next(data)
                print(f'أعمدة Excel: {columns}')
                df = pd.DataFrame(data, columns=columns)
                print('تم تحويل البيانات إلى DataFrame')
            else:
                print('سيتم قراءة الملف كـ CSV')
                df = pd.read_csv(file, encoding='utf-8')
                print('تم تحويل البيانات إلى DataFrame من CSV')
            print(f"أعمدة الملف: {list(df.columns)}")
            print(f"عدد الصفوف: {len(df)}")
            required_cols = [
                'الصفة داخل المكتب', 'الاسم الكامل', 'تاريخ الازدياد', 'مكان الازدياد',
                'المهنة', 'العنوان', 'رقم البطاقة الوطنية', 'اسم الأب', 'اسم الأم', 'الحالة العائلية', 'الجنسية'
            ]
            missing_cols = [col for col in required_cols if col not in df.columns]
            if len(df) == 0:
                flash('الملف لا يحتوي على بيانات!', 'danger')
                print('الملف لا يحتوي على بيانات!')
                return redirect(url_for('dashboard.board'))
            if missing_cols:
                flash(f'الأعمدة التالية مفقودة في الملف: {"، ".join(missing_cols)}', 'danger')
                print(f'الأعمدة التالية مفقودة في الملف: {missing_cols}')
                return redirect(url_for('dashboard.board'))
            for index, row in df.iterrows():
                try:
                    full_name = str(row.get('الاسم الكامل', '')).strip()
                    if ' ' in full_name:
                        parts = full_name.split()
                        first_name = parts[-1]  # الاسم الشخصي هو آخر كلمة
                        last_name = ' '.join(parts[:-1])  # الاسم العائلي هو كل ما قبله
                    else:
                        first_name = full_name
                        last_name = ''
                    birth_date = None
                    if pd.notna(row.get('تاريخ الازدياد', '')):
                        try:
                            birth_date = pd.to_datetime(row.get('تاريخ الازدياد')).date()
                        except:
                            birth_date = None
                    # محاولة ربط فترة الانتداب
                    mandate_title = str(row.get('فترة الانتداب', '')).strip() if 'فترة الانتداب' in row else ''
                    mandate_id = None
                    if mandate_title:
                        mandate = Mandate.query.filter(Mandate.title == mandate_title).first()
                        if mandate:
                            mandate_id = mandate.id
                    member = BoardMember(
                        first_name=first_name,
                        last_name=last_name,
                        position=str(row.get('الصفة داخل المكتب', '')).strip(),
                        birth_date=birth_date,
                        birth_place=str(row.get('مكان الازدياد', '')).strip(),
                        job=str(row.get('المهنة', '')).strip(),
                        address=str(row.get('العنوان', '')).strip(),
                        national_id=str(row.get('رقم البطاقة الوطنية', '')).strip(),
                        father_name=str(row.get('اسم الأب', '')).strip(),
                        mother_name=str(row.get('اسم الأم', '')).strip(),
                        family_status=str(row.get('الحالة العائلية', '')).strip(),
                        nationality=str(row.get('الجنسية', '')).strip(),
                        mandate_period=mandate_title,
                        mandate_id=mandate_id
                    )
                    db.session.add(member)
                    imported_count += 1
                except Exception as e:
                    errors.append(f"خطأ في الصف {index + 2}: {str(e)}")
                    print(f"خطأ في الصف {index + 2}: {str(e)}")
            db.session.commit()
            if errors:
                flash(f'تم استيراد {imported_count} عضو بنجاح. الأخطاء: {', '.join(errors[:5])}', 'warning')
            else:
                flash(f'تم استيراد {imported_count} عضو بنجاح.', 'success')
        except Exception as e:
            flash(f'خطأ في قراءة الملف: {str(e)}', 'danger')
            print('حدث استثناء:', str(e))
        return redirect(url_for('dashboard.board'))
    elif form.validate_on_submit() and current_user.is_admin:
        # التحقق من وجود فترات انتداب قبل الإضافة
        if not mandates:
            flash('لا يمكن إضافة عضو جديد بدون وجود فترات انتداب. يرجى إضافة فترة انتداب أولاً.', 'danger')
            return redirect(url_for('dashboard.board'))
        
        # إضافة عضو جديد
        member = BoardMember(
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            position=form.position.data,
            join_date=form.join_date.data,
            mandate_id=form.mandate_id.data,
            nationality=form.nationality.data,
            family_status=form.family_status.data,
            father_name=form.father_name.data,
            mother_name=form.mother_name.data,
            national_id=form.national_id.data,
            address=form.address.data,
            job=form.job.data,
            birth_place=form.birth_place.data,
            birth_date=form.birth_date.data,
            card_data=form.card_data.data
        )
        if form.image.data:
            image_filename = secure_filename(form.image.data.filename)
            image_path = os.path.join(current_app.root_path, 'static', 'uploads', 'board', image_filename)
            os.makedirs(os.path.dirname(image_path), exist_ok=True)
            form.image.data.save(image_path)
            member.image = f'static/uploads/board/{image_filename}'
        db.session.add(member)
        db.session.commit()
        flash('تم إضافة عضو المكتب بنجاح.', 'success')
        return redirect(url_for('dashboard.board'))
    return render_template('dashboard/board.html', members=members, form=form)

@dashboard_bp.route('/board/edit/<int:member_id>', methods=['GET', 'POST'])
@login_required
def edit_board_member(member_id):
    member = BoardMember.query.get_or_404(member_id)
    mandates = Mandate.query.order_by(Mandate.start_date.desc()).all()
    form = BoardMemberForm(obj=member)
    
    # التحقق من وجود فترات انتداب
    if not mandates:
        flash('يرجى إضافة فترة انتداب أولاً في <a href="{}" class="alert-link">قسم فترات الانتداب</a> قبل تعديل أعضاء المكتب.'.format(url_for('dashboard.mandate')), 'warning')
        form.mandate_id.choices = []
    else:
        form.mandate_id.choices = [(m.id, m.title) for m in mandates]
    
    if form.validate_on_submit() and current_user.is_admin:
        # التحقق من وجود فترات انتداب قبل التعديل
        if not mandates:
            flash('لا يمكن تعديل العضو بدون وجود فترات انتداب. يرجى إضافة فترة انتداب أولاً.', 'danger')
            return redirect(url_for('dashboard.board'))
        
        # Update member data manually instead of using populate_obj
        member.first_name = form.first_name.data
        member.last_name = form.last_name.data
        member.position = form.position.data
        member.join_date = form.join_date.data
        member.mandate_id = form.mandate_id.data
        member.nationality = form.nationality.data
        member.family_status = form.family_status.data
        member.father_name = form.father_name.data
        member.mother_name = form.mother_name.data
        member.national_id = form.national_id.data
        member.address = form.address.data
        member.job = form.job.data
        member.birth_place = form.birth_place.data
        member.birth_date = form.birth_date.data
        member.card_data = form.card_data.data
        
        # Handle image upload separately
        if form.image.data and form.image.data.filename:
            image_filename = secure_filename(form.image.data.filename)
            image_path = os.path.join(current_app.root_path, 'static', 'uploads', 'board', image_filename)
            os.makedirs(os.path.dirname(image_path), exist_ok=True)
            form.image.data.save(image_path)
            member.image = f'static/uploads/board/{image_filename}'
        
        db.session.commit()
        flash('تم تعديل بيانات العضو بنجاح.', 'success')
        return redirect(url_for('dashboard.board'))
    return render_template('dashboard/edit_board_member.html', form=form, member=member)

@dashboard_bp.route('/board/delete/<int:member_id>', methods=['POST'])
@login_required
def delete_board_member(member_id):
    if not current_user.is_admin:
        abort(403)
    member = BoardMember.query.get_or_404(member_id)
    db.session.delete(member)
    db.session.commit()
    flash('تم حذف العضو بنجاح.', 'success')
    return redirect(url_for('dashboard.board'))

@dashboard_bp.route('/board/export', methods=['GET'])
@login_required
def export_board():
    """تصدير أعضاء المكتب إلى ملف Excel"""
    try:
        members = BoardMember.query.all()
        
        # إنشاء DataFrame مع جميع البيانات
        export_data = []
        for member in members:
            export_data.append({
                'first_name': member.first_name,
                'last_name': member.last_name,
                'position': member.position,
                'join_date': member.join_date.strftime('%Y-%m-%d') if member.join_date else '',
                'mandate_period': member.mandate_period,
                'nationality': member.nationality,
                'family_status': member.family_status,
                'father_name': member.father_name,
                'mother_name': member.mother_name,
                'national_id': member.national_id,
                'address': member.address,
                'job': member.job,
                'birth_place': member.birth_place,
                'birth_date': member.birth_date.strftime('%Y-%m-%d') if member.birth_date else '',
                'card_data': member.card_data,
                'created_at': member.created_at.strftime('%Y-%m-%d %H:%M:%S') if member.created_at else '',
                'updated_at': member.updated_at.strftime('%Y-%m-%d %H:%M:%S') if member.updated_at else ''
            })
        
        df = pd.DataFrame(export_data)
        
        # إنشاء ملف Excel في الذاكرة
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='أعضاء المكتب', index=False)
            
            # إضافة ورقة إحصائيات
            stats_data = {
                'الإحصائيات': [
                    f'إجمالي الأعضاء: {len(members)}',
                    f'الرؤساء: {len([m for m in members if m.position == "رئيس"])}',
                    f'نواب الرئيس: {len([m for m in members if m.position == "نائب الرئيس"])}',
                    f'أمناء المال: {len([m for m in members if m.position == "أمين المال"])}',
                    f'تاريخ التصدير: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                ]
            }
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='إحصائيات', index=False)
        
        output.seek(0)
        
        return send_file(
            BytesIO(output.getvalue()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'أعضاء_المكتب_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'danger')
        return redirect(url_for('dashboard.board'))

@dashboard_bp.route('/board/view/<int:member_id>')
@login_required
def view_board_member(member_id):
    member = BoardMember.query.get_or_404(member_id)
    return render_template('dashboard/view_board_member.html', member=member)

@dashboard_bp.route('/board/template', methods=['GET'])
@login_required
def download_board_template():
    """تحميل قالب Excel لاستيراد أعضاء المكتب بالعربية وبتنسيق RTL"""
    try:
        # عناوين الأعمدة بالعربية مرتبة من اليمين لليسار
        columns = [
            'الرقم',
            'الصفة داخل المكتب',
            'الاسم الكامل',
            'تاريخ الازدياد',
            'مكان الازدياد',
            'المهنة',
            'العنوان',
            'رقم البطاقة الوطنية',
            'اسم الأب',
            'اسم الأم',
            'الحالة العائلية',
            'الجنسية'
        ]
        # بيانات تجريبية
        data = [
            [1, 'رئيس', 'محمد أحمد', '1980-01-01', 'الرباط', 'مهندس', 'شارع محمد الخامس، الرباط', 'AB123456', 'أحمد محمد', 'فاطمة علي', 'متزوج', 'مغربية'],
            [2, 'نائب الرئيس', 'فاطمة علي', '1985-05-15', 'الدار البيضاء', 'طبيبة', 'شارع الحسن الثاني، الدار البيضاء', 'CD789012', 'علي حسن', 'خديجة محمد', 'عازبة', 'مغربية']
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'أعضاء المكتب'

        # ضبط اتجاه الورقة من اليمين لليسار
        ws.sheet_view.rightToLeft = True

        # إضافة العناوين
        ws.append(columns)
        # تنسيق العناوين
        header_fill = PatternFill(start_color='B7D3F2', end_color='B7D3F2', fill_type='solid')
        header_font = Font(bold=True)
        for col in range(1, len(columns)+1):
            cell = ws.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            cell.font = header_font

        # إضافة البيانات التجريبية
        for row in data:
            ws.append(row)
        # تنسيق الأعمدة
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # حفظ الملف في الذاكرة
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='قالب_أعضاء_المكتب.xlsx'
        )
    except Exception as e:
        flash(f'خطأ في إنشاء القالب: {str(e)}', 'danger')
        return redirect(url_for('dashboard.board'))

@dashboard_bp.route('/mandate')
@login_required
def mandate():
    mandates = Mandate.query.order_by(Mandate.start_date.desc()).all()
    return render_template('dashboard/mandate.html', mandates=mandates)

@dashboard_bp.route('/mandate/add', methods=['GET', 'POST'])
@login_required
def add_mandate():
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('dashboard.mandate'))
    
    form = MandateForm()
    if form.validate_on_submit():
        mandate = Mandate(
            title=form.title.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            description=form.description.data
        )
        db.session.add(mandate)
        db.session.commit()
        flash('تم إضافة فترة الانتداب بنجاح', 'success')
        return redirect(url_for('dashboard.mandate'))
    
    return render_template('dashboard/forms/mandate_form.html', form=form)

@dashboard_bp.route('/mandate/<int:mandate_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_mandate(mandate_id):
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('dashboard.mandate'))
    
    mandate = Mandate.query.get_or_404(mandate_id)
    form = MandateForm()
    
    if form.validate_on_submit():
        mandate.title = form.title.data
        mandate.start_date = form.start_date.data
        mandate.end_date = form.end_date.data
        mandate.description = form.description.data
        db.session.commit()
        flash('تم تحديث فترة الانتداب بنجاح', 'success')
        return redirect(url_for('dashboard.mandate'))
    
    form.title.data = mandate.title
    form.start_date.data = mandate.start_date
    form.end_date.data = mandate.end_date
    form.description.data = mandate.description
    
    return render_template('dashboard/forms/mandate_form.html', form=form, mandate=mandate)

@dashboard_bp.route('/mandate/<int:mandate_id>/delete', methods=['POST'])
@login_required
def delete_mandate(mandate_id):
    if not current_user.is_admin:
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('dashboard.mandate'))
    
    mandate = Mandate.query.get_or_404(mandate_id)
    db.session.delete(mandate)
    db.session.commit()
    flash('تم حذف فترة الانتداب بنجاح', 'success')
    return redirect(url_for('dashboard.mandate'))

@dashboard_bp.route('/assets')
@login_required
def assets():
    return render_template('dashboard/assets.html', title='ممتلكات الجمعية') 